from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors,Font
import locale,time
from Config.ProjConfigVar import *
from Util.Log import logger

class Excel(object):

    def __init__(self,excel_file_path):
        self.excle_file_path = excel_file_path
        self.wb = load_workbook(excel_file_path)
        self.ws = self.wb[self.wb.sheetnames[0]]

    #获取所有sheet名称
    def get_all_sheet_names(self):
        return self.wb.sheetnames

    #获取指定位置的sheet名称
    def get_sheet_name_by_index(self,index):
        return self.wb.sheetnames[index - 1]

    #获取excel表格的路径
    def get_excel_file_path(self):
        return self.excle_file_path

    #创建sheet
    def create_sheet1(self,sheet_name,position = None):
        try:
            if position:
                self.wb.create_sheet(sheet_name,position)
            else:
                self.wb.create_sheet(sheet_name)
            return True
        except Exception as e:
            #print(e)
            logger.debug(e)
            return False

    #设置当前要操作的sheet名称是哪个
    def get_sheet_by_name(self,sheet_name):
        if sheet_name not in self.wb.sheetnames:
            #print("%s sheet不存在，请重新设置！"%sheet_name)
            logger.warning("%s sheet不存在，请重新设置！"%sheet_name)
            return False
        self.ws = self.wb[sheet_name]
        return True

    #设定指定的sheet名称
    def get_sheet_by_index(self,index):
        self.ws = self.wb[self.get_sheet_name_by_index(index)]
        #print("设定的sheet名称是：",self.ws.title)
        logger.info("设定的sheet名称是：%s"% self.ws.title)

    #读指定的某个单元格中的值
    def get_cell_value(self,row_no,col_no,sheet_name=None):
        if sheet_name is not None:
            result = self.get_sheet_by_name(sheet_name)
            if result ==False:
                return None
        return self.ws.cell(row_no,col_no).value

    #读某一行的值
    def get_single_row_values(self,row_no,sheet_name=None):
        cell_values = []
        if sheet_name is not None:
            result = self.get_sheet_by_name(sheet_name)
            if result ==False:
                return None
        for cell in list(self.ws.rows)[row_no - 1]:
            cell_values.append(cell.value)
        return cell_values

    #读取某个sheet的所有行中的单元格内容，使用2维的列表进行存储
    def get_all_row_values(self,sheet_name=None):
        all_cell_values = []    #所有的单元格的值均存入列表
        if sheet_name is not None:    #参数设置了新的sheet
            result = self.get_sheet_by_name(sheet_name)
            if result ==False:
                return None
        for row in list(self.ws.rows):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            all_cell_values.append(row_values)
        return all_cell_values

    #读某一列的值
    def get_single_col_values(self,col_no,sheet_name=None):
        cell_values = []
        if sheet_name is not None:
            result = self.get_sheet_by_name(sheet_name)
            if result ==False:
                return None
        for cell in list(self.ws.columns)[col_no - 1]:
            cell_values.append(cell.value)
        return cell_values

    #读某个范围的值
    def get_some_values(self,min_row_no,min_col_no,max_row_no,max_col_no,sheet_name=None):
        if sheet_name is not None:    #参数化设置了新的sheet
            result = self.get_sheet_by_name(sheet_name)
            if result==False:
                return None
        values =[]
        for i in range(min_row_no,max_row_no+1):
            row_values = []
            for j in range(min_col_no,max_col_no+1):
                row_values.append(self.ws.cell(row=i,column=j).value)
            values.append(row_values)
        return values

    #保存
    def save(self):
        self.wb.save(self.excle_file_path)

    #写
    def write_cell_value(self,row_no,col_no,value,style=None,sheet_name=None):
        if sheet_name is not None:
            result = self.get_sheet_by_name(sheet_name)
            if result ==False:
                return None
        if style is None:
            style = colors.BLACK
        elif style == 'red':
            style = colors.RED
        elif style =='green':
            style = colors.GREEN
        self.ws.cell(row=row_no,column=col_no).font = Font(color=style)
        self.ws.cell(row=row_no,column=col_no,value=value)
        self.save()
        return True

    #写时间
    def write_current_time(self,row_no,col_no,style=None,sheet_name=None):
        if sheet_name is not None:
            result = self.get_sheet_by_name(sheet_name)
            if result==False:
                return None
        if style is None:
            style = colors.BLACK
        elif style =='red':
            style = colors.RED
        elif style =='green':
            style = colors.GREEN
        locale.setlocale(locale.LC_ALL,'en')
        locale.setlocale(locale.LC_CTYPE,'chinese')
        self.ws.cell(row=row_no,column=col_no).font = Font(color=style)
        self.ws.cell(row=row_no, column=col_no, value=time.strftime('%Y年%m月%d日 %H时%M分%S秒'))
        self.save()
        return True

if __name__=="__main__":
    import os
    from Config.interfaceInfo import proj_path
    #获取测试数据文件的绝对路径
    test_data_excel_path = os.path.join(proj_path,"TestData\\接口测试数据.xlsx")
    print(test_data_excel_path)

    test_cases_wb = Excel(test_data_excel_path)
    test_data_sheet_name = "测试用例集"
    test_cases_wb.get_sheet_by_name(test_data_sheet_name)
    print("设定的测试用例sheet名称为：%s"%test_data_sheet_name)
    print("*"*50)

    #print("/"*50,test_cases_wb.get_sheet_by_index())

    #读取所有的接口测试用例
    print("test_cases_wb.get_all_row_values():%s"%test_cases_wb.get_all_row_values())
    test_cases = []
    for row in test_cases_wb.get_all_row_values()[1:]:
        test_case = row[test_data_interface_name_col_no], row[test_data_request_data_col_no], row[
            test_data_assert_word_col_no]
        test_cases.append(test_case)

    print(test_cases)





